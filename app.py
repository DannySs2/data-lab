from dotenv import load_dotenv
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from src.get_items_id import create_dataframe


load_dotenv()

organization = "horbath"
project = "Promigas"
pat = os.getenv("TOKEN")

df_ids = pd.read_csv("data/ids.csv")

itemsID = df_ids["ID"].dropna().astype(int).tolist()

fecha_corte = "2025-03-01T05:00:00Z"

df_items = create_dataframe(
    itemsID,
    organization,
    project,
    pat,
    fecha_corte
)


output_path = "./data/resultado.xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_items.to_excel(writer, index=False, sheet_name="Resultado")
    ws = writer.sheets["Resultado"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

print(f"Archivo {output_path} generado con {len(df_items)} filas")