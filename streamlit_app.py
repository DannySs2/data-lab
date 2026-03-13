import streamlit as st
import pandas as pd
import io
from datetime import date
from dotenv import load_dotenv
import os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from src.get_items_id import create_dataframe

load_dotenv()

organization = "horbath"
project = "Promigas"
pat = os.getenv("TOKEN")

st.title("Reporte de estados - Azure DevOps")

# --- inputs ---
fecha_corte_date = st.date_input(
    "Fecha de corte",
    value=date.today()
)

fecha_corte = f"{fecha_corte_date.isoformat()}T00:00:00Z"
st.info(f"Se filtrará el estado activo al: {fecha_corte}")

ids_file = st.file_uploader("Sube el archivo ids.csv", type=["csv"])

if st.button("Generar reporte"):
    if ids_file is None:
        st.error("Por favor sube el archivo ids.csv")
    else:
        df_ids = pd.read_csv(ids_file)
        itemsID = df_ids["ID"].dropna().astype(int).tolist()

        with st.spinner(f"Extrayendo datos de {len(itemsID)} items..."):
            df_resultado = create_dataframe(
                itemsID,
                organization,
                project,
                pat,
                fecha_corte
            )

        st.session_state["df_resultado"] = df_resultado
        st.success(f"Reporte generado con {len(df_resultado)} filas")

if "df_resultado" in st.session_state:
    df_resultado = st.session_state["df_resultado"]
    st.dataframe(df_resultado)

    # --- generar Excel en memoria ---
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_resultado.to_excel(writer, index=False, sheet_name="Resultado")
        ws = writer.sheets["Resultado"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    buffer.seek(0)
    st.download_button(
        label="Descargar Excel",
        data=buffer,
        file_name="resultado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
